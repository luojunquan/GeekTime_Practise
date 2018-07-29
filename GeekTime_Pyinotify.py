#!/usr/bin/env python
#encoding:UTF-8
# import pyinotify
'''
wm = pyinotify.WatchManager()
#监视创建和删除事件
mask = pyinotify.IN_DELETE | pyinotify.IN_CREATE
#使用 WatchManager对象的 add_watch 方法添加对文件的监视事件
wm.add_watch('/tmp',mask)

notifier = pyinotify.Notifier(wm)

print(notifier.loop())
'''
##定制化事件处理器
import glob

'''
vm = pyinotify.WatchManager()
mask = pyinotify.IN_DELETE | pyinotify.IN_CREATE
#定制事件处理方式的方法是，继承 ProcessEvent 类
class EventHandler(pyinotify.ProcessEvent):
    def process_IN_CREATE(self,event):
        print("Creating",event.pathname)

    def process_IN_DELETE(self,event):
        print("Removeing", event.pathname)

handler = EventHandler()
notifier = pyinotify.Notifier(vm,handler)
wd = vm.add_watch('/tmp',mask,rec=True)

notifier.loop()
'''
#########################
import openpyxl
wb = openpyxl.load_workbook('PM1.xlsx')
# print(wb.active)
# print(wb.read_only)
# print(wb.encoding)
# print(wb.worksheets)
print(wb.sheetnames)
#python3改成如下内容：wb.get_sheet_by_name('统计')
ws = wb['路由策略']
# print(ws.title)
# print(ws.dimensions)
# print(ws['B2'])
# print(ws.cell(row=3,column=4))
# for row in ws.values:
#     print(*row)
'''
for row in ws.rows:
    print(*[cell.value for cell in row])
    print('\n')
'''
#上下两种方式的效果是一样的
'''
for i in range(ws.min_row,ws.max_row + 1):
    for j in range(ws.min_column,ws.max_column + 1):
        print(ws.cell(row=i,column=j).value ,end=' ')
    print('\n')
'''
#########################################################
#计算出总分和平均成绩
'''
def process_worksheet(sheet):
    avg_column = sheet.max_column + 1
    sum_column = sheet.max_column + 2

# 通过传递 min row 表示从第二行开始遍历 表格的第一列保存的是学生的学号，第二列保存的是学生的姓名，这也是我们
# 在计算平均分和总分时不会使用的数据 因此，我们通过传递 min_col为3表示从第三列开始遍历,iter_rows 函数按行返回单元格，
# 因此，我们只需要循环遍历 iter_rows 函数的结果，就实现了计算每一位学生的平均分和总分的功能
# iter_rows 函数返回的是 Cell 对象，所以，我们在计算成绩之前需要先通过一个列表推导表达式，得到每一个单元格的取值

    for row in sheet.iter_rows(min_row = 2,min_col = 3):
        scores = [cell.value for cell in row]
        sum_score = sum(scores)
        avg_score = sum_score / len(scores)
        #计算平均分和总分，并且保存到最后两列
        sheet.cell(row = row[0].row,column  = avg_column).value = avg_score
        sheet.cell(row = row[0].row,column = sum_column).value = sum_score
    #设置平均分和总分的标题部分
    sheet.cell(row = 1,column = avg_column).value = 'avg'
    sheet.cell(row = 1,column = sum_column).value = 'sum'


def main():
    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb['student']
    process_worksheet(sheet)
    wb.save('example_copy.xlsx')

if __name__ == '__main__':
    main()
 '''

###########################################
##合并多个excel表
'''
import os

def get_all_xlsx_files(path):
    #通过global获取指定目录下的所有以p开头的所有xlsx文档
    xlsx_files = glob.glob(os.path.join(path,'p*.xlsx'))
    #按照文件名称进行排序
    sorted(xlsx_files,key = str.lower)
    return xlsx_files

def merge_xlsx_files(xlsx_files):
    wb = openpyxl.load_workbook(xlsx_files[0])
    ws = wb.active
    #命名合并表的名称
    ws.title = 'merge result'

    for filename in xlsx_files[1:]:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        for row  in sheet.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            ws.append(values)
    return wb


def main():
    xlsx_files = get_all_xlsx_files(os.path.expanduser('./'))
    wb = merge_xlsx_files(xlsx_files)
    wb.save('merge_form.xlsx')

if __name__ == '__main__':
    main()
'''
###########################




























